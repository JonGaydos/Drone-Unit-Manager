import io
from datetime import datetime, date
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from app.models.pilot import Pilot
from app.models.vehicle import Vehicle
from app.models.flight import Flight, FlightPurpose
from app.models.certification import CertificationType, PilotCertification, PilotEquipmentQual


def _parse_date(val) -> Optional[date]:
    if val is None or val == "Pending" or val == "N/A" or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _parse_status(val) -> str:
    if val is None or val == "" or val == "Pending":
        return "pending"
    if isinstance(val, datetime) or isinstance(val, date):
        return "complete"
    s = str(val).strip().lower()
    status_map = {
        "complete": "complete",
        "issued": "complete",
        "active": "active",
        "in progress": "in_progress",
        "pending": "pending",
        "not started": "not_started",
        "not issued": "not_issued",
        "not eligible": "not_eligible",
        "n/a": "not_eligible",
        "exempt": "complete",
    }
    return status_map.get(s, "pending")


def _find_or_create_pilot(db: Session, full_name: str) -> tuple[Optional[Pilot], bool]:
    """Returns (pilot, was_created)."""
    if not full_name or full_name.strip() == "":
        return None, False
    parts = full_name.strip().split(" ", 1)
    first = parts[0]
    last = parts[1] if len(parts) > 1 else ""
    pilot = db.query(Pilot).filter(Pilot.first_name == first, Pilot.last_name == last).first()
    if not pilot:
        pilot = Pilot(first_name=first, last_name=last, status="active")
        db.add(pilot)
        db.flush()
        return pilot, True
    return pilot, False


def _find_or_create_vehicle(db: Session, vehicle_str: str) -> tuple[Optional[Vehicle], bool]:
    """Returns (vehicle, was_created)."""
    if not vehicle_str or vehicle_str.strip() == "":
        return None, False
    serial = vehicle_str.strip()
    vehicle = db.query(Vehicle).filter(Vehicle.serial_number == serial).first()
    if not vehicle:
        manufacturer = "Skydio"
        model = serial
        if "X10" in serial:
            model = "X10"
        elif "X2" in serial:
            model = "X2E"
        elif "BRINC" in serial or "LEMUR" in serial:
            manufacturer = "BRINC"
            model = "LEMUR 2"
        vehicle = Vehicle(
            serial_number=serial,
            manufacturer=manufacturer,
            model=model,
            status="active",
        )
        db.add(vehicle)
        db.flush()
        return vehicle, True
    return vehicle, False


def _ensure_cert_type(db: Session, name: str, category: str, has_expiration: bool = True,
                      renewal_months: Optional[int] = None) -> CertificationType:
    ct = db.query(CertificationType).filter(CertificationType.name == name).first()
    if not ct:
        ct = CertificationType(
            name=name, category=category, has_expiration=has_expiration,
            renewal_period_months=renewal_months, is_active=True,
        )
        db.add(ct)
        db.flush()
    return ct


def _ensure_purpose(db: Session, name: str):
    if not name or name.strip() == "":
        return
    existing = db.query(FlightPurpose).filter(FlightPurpose.name == name.strip()).first()
    if not existing:
        db.add(FlightPurpose(name=name.strip(), sort_order=100))
        db.flush()


def _parse_safe_float(val) -> Optional[float]:
    """Safely parse a float from a cell value."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _is_duplicate_flight(db: Session, flight_id: str, pilot, vehicle, flight_date, duration) -> bool:
    """Check if a flight is a duplicate by external_id or by pilot+date+duration+vehicle."""
    from sqlalchemy import func as sa_func
    existing = db.query(Flight).filter(
        sa_func.replace(sa_func.upper(Flight.external_id), "-", "") == flight_id
    ).first()
    if existing:
        return True

    if flight_date and duration:
        tolerance = int(duration * 0.1) or 30
        potential = db.query(Flight).filter(
            Flight.date == flight_date,
            Flight.pilot_id == (pilot.id if pilot else None),
            Flight.vehicle_id == (vehicle.id if vehicle else None),
            Flight.duration_seconds.between(duration - tolerance, duration + tolerance),
        ).first()
        if potential:
            return True

    return False


def _import_skydio_sheet(ws, db: Session, result: dict) -> tuple[int, int]:
    """Import flights from the Skydio sheet. Returns (pilots_created, vehicles_created)."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    pilots_created = 0
    vehicles_created = 0

    for r in range(2, ws.max_row + 1):
        try:
            row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
            flight_id = row.get("Flight ID")
            if not flight_id:
                continue
            flight_id = str(flight_id).upper().replace("-", "")

            pilot, pilot_new = _find_or_create_pilot(db, row.get("Pilot", ""))
            if pilot_new:
                pilots_created += 1
            vehicle, vehicle_new = _find_or_create_vehicle(db, row.get("Vehicle", ""))
            if vehicle_new:
                vehicles_created += 1

            takeoff_time_raw = row.get("Takeoff") or row.get("Local Takeoff Time")
            flight_date = takeoff_time_raw.date() if isinstance(takeoff_time_raw, datetime) else None
            duration = row.get("Duration (seconds)")
            duration = int(duration) if duration else None

            if _is_duplicate_flight(db, flight_id, pilot, vehicle, flight_date, duration):
                result["flights_skipped"] += 1
                continue

            purpose = row.get("Purpose", "")
            if purpose:
                _ensure_purpose(db, purpose)

            lat = _parse_safe_float(row.get("Takeoff Latitude"))
            lon = _parse_safe_float(row.get("Takeoff Longitude"))

            flight = Flight(
                external_id=str(flight_id),
                api_provider="excel_import",
                pilot_id=pilot.id if pilot else None,
                vehicle_id=vehicle.id if vehicle else None,
                date=flight_date,
                takeoff_time=takeoff_time_raw if isinstance(takeoff_time_raw, datetime) else None,
                landing_time=row.get("Land") if isinstance(row.get("Land"), datetime) else None,
                duration_seconds=int(duration) if duration else None,
                takeoff_lat=lat,
                takeoff_lon=lon,
                takeoff_address=row.get("Takeoff Address", ""),
                purpose=purpose if purpose else None,
                battery_serial=row.get("Battery", ""),
                sensor_package=row.get("Sensor Package", "") or None,
                attachment_top=row.get("Attachment (TOP)", "") or None,
                attachment_bottom=row.get("Attachment (BOTTOM)", "") or None,
                attachment_left=row.get("Attachment (LEFT)", "") or None,
                attachment_right=row.get("Attachment (RIGHT)", "") or None,
                carrier=row.get("Carrier(s)", "") or None,
                review_status="reviewed",
                pilot_confirmed=True,
                data_source="excel_import",
            )
            db.add(flight)
            try:
                from app.services.sync_manager import _ensure_equipment_records
                _ensure_equipment_records(db, flight)
            except Exception:
                pass
            result["flights_imported"] += 1
        except Exception as e:
            result["errors"].append(f"Skydio row {r}: {str(e)}")

    return pilots_created, vehicles_created


# Certification definitions and column mapping for Pilot Info sheet
_CERT_DEFS = [
    ("Skydio X2E Academy", "equipment", False, None),
    ("Skydio X2E Checkoff", "equipment", False, None),
    ("Skydio X10 Academy", "equipment", False, None),
    ("Skydio X10 Checkoff", "equipment", False, None),
    ("BRINC LEMUR 2 Academy", "equipment", False, None),
    ("BRINC LEMUR 2 Checkoff", "equipment", False, None),
    ("NIST Level 1", "nist", False, None),
    ("NIST Level 2", "nist", False, None),
    ("NIST Level 3", "nist", False, None),
    ("NIST Level 4", "nist", False, None),
    ("NIST Level 5", "nist", False, None),
    ("DART Training", "training", False, None),
    ("GCSC Training", "training", False, None),
    ("Insurance 2026", "insurance", True, 12),
    ("Insurance 2027", "insurance", True, 12),
    ("FAA Part 107", "faa", True, 24),
]

_COL_MAP = {
    3: "Skydio X2E Academy", 4: "Skydio X2E Checkoff",
    5: "Skydio X10 Academy", 6: "Skydio X10 Checkoff",
    7: "BRINC LEMUR 2 Academy", 8: "BRINC LEMUR 2 Checkoff",
    9: "NIST Level 1", 10: "NIST Level 2", 11: "NIST Level 3",
    12: "NIST Level 4", 13: "NIST Level 5",
    14: "DART Training", 15: "GCSC Training",
    16: "Insurance 2026", 17: "Insurance 2027",
}


def _assign_cert_columns(db: Session, pilot: Pilot, ws, r: int, col_map: dict, cert_types: dict, result: dict):
    """Assign certifications from column-mapped cells for a single pilot row."""
    for col, cert_name in col_map.items():
        val = ws.cell(r, col).value
        ct = cert_types.get(cert_name)
        if not ct:
            continue

        existing = db.query(PilotCertification).filter(
            PilotCertification.pilot_id == pilot.id,
            PilotCertification.certification_type_id == ct.id,
        ).first()
        if existing:
            continue

        pc = PilotCertification(
            pilot_id=pilot.id,
            certification_type_id=ct.id,
            status=_parse_status(val),
            issue_date=_parse_date(val),
        )
        db.add(pc)
        result["cert_assignments"] += 1


def _assign_faa_part107(db: Session, pilot: Pilot, ws, r: int, cert_types: dict, result: dict):
    """Assign FAA Part 107 certification for a pilot row."""
    ct_faa = cert_types.get("FAA Part 107")
    if not ct_faa:
        return

    existing = db.query(PilotCertification).filter(
        PilotCertification.pilot_id == pilot.id,
        PilotCertification.certification_type_id == ct_faa.id,
    ).first()
    if existing:
        return

    part107_date = _parse_date(ws.cell(r, 19).value)
    renewal_due = _parse_date(ws.cell(r, 20).value)
    renewed_date = _parse_date(ws.cell(r, 21).value)
    renewal_due_2 = _parse_date(ws.cell(r, 22).value)

    faa_status = "active" if part107_date else "pending"
    if renewal_due and not renewed_date and renewal_due < date.today():
        faa_status = "expired"

    pc = PilotCertification(
        pilot_id=pilot.id,
        certification_type_id=ct_faa.id,
        status=faa_status,
        issue_date=renewed_date or part107_date,
        expiration_date=renewal_due_2 or renewal_due,
    )
    db.add(pc)
    result["cert_assignments"] += 1


def _import_pilot_info_sheet(ws, db: Session, result: dict) -> int:
    """Import pilot certifications from the Pilot Info sheet. Returns pilots_created count."""
    cert_types = {}
    for name, cat, has_exp, months in _CERT_DEFS:
        ct = _ensure_cert_type(db, name, cat, has_exp, months)
        cert_types[name] = ct
        result["certifications_created"] += 1

    pilots_created = 0

    for r in range(7, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or str(name).strip() == "" or "STATUS KEY" in str(name):
            continue

        try:
            pilot, pilot_new = _find_or_create_pilot(db, str(name))
            if pilot_new:
                pilots_created += 1
            if not pilot:
                continue

            status_val = ws.cell(r, 2).value
            if status_val:
                pilot.status = "active" if str(status_val).lower() == "active" else "inactive"

            _assign_cert_columns(db, pilot, ws, r, _COL_MAP, cert_types, result)
            _assign_faa_part107(db, pilot, ws, r, cert_types, result)

        except Exception as e:
            result["errors"].append(f"Pilot Info row {r}: {str(e)}")

    return pilots_created


def import_excel(db: Session, file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    pilots_created_count = 0
    vehicles_created_count = 0
    result = {
        "pilots_created": 0,
        "vehicles_created": 0,
        "flights_imported": 0,
        "flights_skipped": 0,
        "certifications_created": 0,
        "cert_assignments": 0,
        "errors": [],
    }

    if "Skydio" in wb.sheetnames:
        p_count, v_count = _import_skydio_sheet(wb["Skydio"], db, result)
        pilots_created_count += p_count
        vehicles_created_count += v_count

    if "Pilot Info" in wb.sheetnames:
        pilots_created_count += _import_pilot_info_sheet(wb["Pilot Info"], db, result)

    result["pilots_created"] = pilots_created_count
    result["vehicles_created"] = vehicles_created_count

    db.commit()
    return result
